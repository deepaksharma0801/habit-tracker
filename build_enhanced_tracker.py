from __future__ import annotations

from copy import copy
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, RadarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import Outline
from openpyxl.workbook.defined_name import DefinedName


ROOT = Path("/Users/snadimi3/Documents/habit tracker")
OUTPUT_XLSX = ROOT / "SerenatasHabitTracker_Enhanced_Dashboard.xlsx"
APPS_SCRIPT_FILE = ROOT / "habit_dashboard.gs"

COLORS = {
    "green": "6BBF77",
    "green_dark": "2E7D57",
    "green_soft": "DFF6E0",
    "amber": "D9A441",
    "amber_soft": "F7E8C3",
    "red": "D96C6C",
    "red_soft": "F9D8D6",
    "gray_bg": "EEF2F4",
    "gray_alt": "F7F9FA",
    "gray_line": "D6DEE2",
    "text": "24343D",
    "white": "FFFFFF",
    "gold": "E7C66E",
    "teal": "7AA5A9",
}

HABITS = [
    ("Morning Exercise", "Health"),
    ("Hydration 2L", "Health"),
    ("Sleep by 10 pm", "Health"),
    ("No Junk Food", "Health"),
    ("Walk 10k Steps", "Health"),
    ("Cook at Home", "Health"),
    ("Read 30 min", "Learning"),
    ("Journal", "Learning"),
    ("Learn a Language", "Learning"),
    ("Gratitude List", "Learning"),
    ("Practice a Skill", "Learning"),
    ("Creative Work", "Learning"),
    ("Learning Module", "Learning"),
    ("Deep Work Block", "Career"),
    ("Review Goals", "Career"),
    ("No Social Media", "Career"),
    ("Networking Task", "Career"),
    ("Side Project", "Career"),
    ("Read Industry News", "Career"),
    ("Email Zero", "Career"),
    ("Plan Tomorrow", "Career"),
    ("Weekly Review", "Career"),
    ("Savings Transfer", "Finance"),
    ("Track Spending", "Finance"),
    ("No Impulse Buys", "Finance"),
    ("Affirmations", "Mindset"),
    ("Outdoor Time", "Mindset"),
    ("Digital Detox", "Mindset"),
    ("Meditation 10 min", "Mindset"),
    ("Portfolio Update", "Career"),
]

APP_SCRIPT = """function onOpen() {
  SpreadsheetApp.getUi()
    .createMenu('Tracker Tools')
    .addItem('Setup / Refresh dashboard', 'setupHabitDashboard')
    .addItem('Archive current month', 'archiveCurrentMonth')
    .addToUi();
  ensureNamedRanges_();
}

function onEdit(e) {
  if (!e || !e.range) return;
  const sheet = e.range.getSheet();
  if (sheet.getName() !== 'Tracker') return;
  if (e.range.getA1Notation() === 'D1') {
    updateTrackerMonth_();
    ensureNamedRanges_();
    buildDashboardCharts_();
  }
}

function setupHabitDashboard() {
  updateTrackerMonth_();
  applyTrackerCheckboxes_();
  ensureNamedRanges_();
  buildDashboardCharts_();
}

function updateTrackerMonth_() {
  const ss = SpreadsheetApp.getActive();
  const tracker = ss.getSheetByName('Tracker');
  const dashboard = ss.getSheetByName('Dashboard');
  const monthValue = tracker.getRange('D1').getValue();
  const monthStart = new Date(monthValue);
  if (!monthValue || Number.isNaN(monthStart.getTime())) return;

  monthStart.setDate(1);
  tracker.getRange('D1').setValue(monthStart).setNumberFormat('mmmm yyyy');
  const daysInMonth = new Date(monthStart.getFullYear(), monthStart.getMonth() + 1, 0).getDate();

  const headerValues = [];
  for (let i = 1; i <= 31; i += 1) {
    if (i <= daysInMonth) {
      headerValues.push([new Date(monthStart.getFullYear(), monthStart.getMonth(), i)]);
    } else {
      headerValues.push(['']);
    }
  }

  tracker.getRange(2, 4, 1, 31).setValues([headerValues.map(v => v[0])]).setNumberFormat('d');
  tracker.showColumns(4, 31);
  if (daysInMonth < 31) tracker.hideColumns(4 + daysInMonth, 31 - daysInMonth);

  applyTrackerCheckboxes_();
  dashboard.getRange('A1').setFormula('="Monthly Habit Dashboard — "&TEXT(Tracker!D1,"MMMM YYYY")');
}

function applyTrackerCheckboxes_() {
  const tracker = SpreadsheetApp.getActive().getSheetByName('Tracker');
  const range = tracker.getRange('D3:AH32');
  range.insertCheckboxes();
}

function ensureNamedRanges_() {
  const ss = SpreadsheetApp.getActive();
  const tracker = ss.getSheetByName('Tracker');
  const summary = ss.getSheetByName('Summary');
  const dashboard = ss.getSheetByName('Dashboard');

  const namedRanges = [
    ['Tracker_Month', tracker.getRange('D1')],
    ['Tracker_Dates', tracker.getRange('D2:AH2')],
    ['Tracker_Habits', tracker.getRange('B3:B32')],
    ['Tracker_Categories', tracker.getRange('C3:C32')],
    ['Tracker_DataRange', tracker.getRange('D3:AH32')],
    ['Summary_Progress', summary.getRange('D2:D31')],
    ['Dashboard_TopN', dashboard.getRange('Q11')],
  ];

  namedRanges.forEach(([name, range]) => {
    const existing = ss.getNamedRanges().find(item => item.getName() === name);
    if (existing) {
      existing.remove();
    }
    ss.setNamedRange(name, range);
  });
}

function archiveCurrentMonth() {
  const ss = SpreadsheetApp.getActive();
  const tracker = ss.getSheetByName('Tracker');
  const summary = ss.getSheetByName('Summary');
  const history = ss.getSheetByName('History');
  const monthStart = tracker.getRange('D1').getValue();
  if (!monthStart) return;

  const existing = history.getDataRange().getValues();
  const stamp = new Date(monthStart);
  stamp.setDate(1);
  const stampKey = Utilities.formatDate(stamp, Session.getScriptTimeZone(), 'yyyy-MM-dd');

  const rows = summary.getRange('B2:F31').getValues()
    .filter(row => row[0])
    .map(row => [stamp, row[0], row[1], row[2], row[3], row[4], new Date()]);

  const alreadyArchived = existing.some((row, index) => {
    if (index === 0) return false;
    return Utilities.formatDate(new Date(row[0]), Session.getScriptTimeZone(), 'yyyy-MM-dd') === stampKey;
  });

  if (!alreadyArchived && rows.length) {
    history.getRange(history.getLastRow() + 1, 1, rows.length, rows[0].length).setValues(rows);
  }
}

function buildDashboardCharts_() {
  const ss = SpreadsheetApp.getActive();
  const dashboard = ss.getSheetByName('Dashboard');
  const helper = ss.getSheetByName('Helper');

  dashboard.getCharts().forEach(chart => dashboard.removeChart(chart));

  const lineChart = dashboard.newChart()
    .asLineChart()
    .addRange(helper.getRange('B1:D32'))
    .setPosition(10, 1, 0, 0)
    .setNumHeaders(1)
    .setOption('title', 'Monthly Completion Trend')
    .setOption('curveType', 'function')
    .setOption('legend.position', 'bottom')
    .setOption('pointSize', 5)
    .setOption('hAxis.title', 'Day')
    .setOption('vAxis.title', 'Completion %')
    .setOption('vAxis.viewWindow.min', 0)
    .setOption('vAxis.viewWindow.max', 1)
    .setOption('chartArea', {left: 60, top: 40, width: '78%', height: '70%'})
    .setOption('series', {
      0: {color: '#2E7D57', lineWidth: 3},
      1: {color: '#7AA5A9', lineDashStyle: [6, 4], lineWidth: 2, targetAxisIndex: 1}
    })
    .build();
  dashboard.insertChart(lineChart);

  const barChart = dashboard.newChart()
    .asBarChart()
    .addRange(helper.getRange('O1:P21'))
    .setPosition(38, 1, 0, 0)
    .setNumHeaders(1)
    .setOption('title', 'Habit Consistency (Top N)')
    .setOption('legend.position', 'none')
    .setOption('hAxis.title', 'Completion %')
    .setOption('vAxis.title', 'Habit')
    .setOption('hAxis.viewWindow.min', 0)
    .setOption('hAxis.viewWindow.max', 1)
    .setOption('annotations.alwaysOutside', true)
    .setOption('series', {0: {color: '#6BBF77'}})
    .build();
  dashboard.insertChart(barChart);

  const streakChart = dashboard.newChart()
    .asColumnChart()
    .addRange(helper.getRange('R1:T11'))
    .setPosition(24, 9, 0, 0)
    .setNumHeaders(1)
    .setOption('title', 'Streak Leaderboard')
    .setOption('legend.position', 'bottom')
    .setOption('hAxis.title', 'Habit')
    .setOption('vAxis.title', 'Days')
    .setOption('series', {
      0: {color: '#2E7D57'},
      1: {color: '#D9A441'}
    })
    .build();
  dashboard.insertChart(streakChart);

  const categoryChart = dashboard.newChart()
    .asColumnChart()
    .addRange(helper.getRange('U1:V10'))
    .setPosition(38, 9, 0, 0)
    .setNumHeaders(1)
    .setOption('title', 'Category Completion')
    .setOption('legend.position', 'none')
    .setOption('hAxis.title', 'Category')
    .setOption('vAxis.title', 'Avg Completion %')
    .setOption('vAxis.viewWindow.min', 0)
    .setOption('vAxis.viewWindow.max', 1)
    .setOption('series', {0: {color: '#7AA5A9'}})
    .build();
  dashboard.insertChart(categoryChart);

  const donutChart = dashboard.newChart()
    .asPieChart()
    .addRange(helper.getRange('W1:X3'))
    .setPosition(4, 16, 0, 0)
    .setNumHeaders(1)
    .setOption('title', 'Monthly Progress')
    .setOption('pieHole', 0.68)
    .setOption('legend.position', 'none')
    .setOption('pieSliceText', 'percentage')
    .setOption('slices', {
      0: {color: '#2E7D57'},
      1: {color: '#D6DEE2'}
    })
    .build();
  dashboard.insertChart(donutChart);
}
"""


def hex_fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def apply_range_fill(ws, cell_range: str, color: str) -> None:
    fill = hex_fill(color)
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.fill = copy(fill)


def set_all_borders(ws, cell_range: str, color: str = COLORS["gray_line"], style: str = "thin") -> None:
    side = Side(style=style, color=color)
    border = Border(left=side, right=side, top=side, bottom=side)
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border


def add_named_range(wb: Workbook, name: str, ref: str) -> None:
    wb.defined_names.add(DefinedName(name, attr_text=ref))


def style_sheet_defaults(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.outlinePr = Outline(summaryBelow=False, summaryRight=False)


def style_title(ws, cell: str, text: str, end_cell: str) -> None:
    ws.merge_cells(f"{cell}:{end_cell}")
    c = ws[cell]
    c.value = text
    c.font = Font(name="Montserrat", size=20, bold=True, color=COLORS["text"])
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.fill = hex_fill(COLORS["gray_bg"])
    ws.row_dimensions[c.row].height = 28
    set_all_borders(ws, f"{cell}:{end_cell}")


def build_habits(ws) -> None:
    style_sheet_defaults(ws)
    ws.title = "Habits"
    ws.freeze_panes = "A2"
    headers = ["No.", "Habit Name", "Category"]
    for idx, header in enumerate(headers, 1):
        cell = ws.cell(1, idx, header)
        cell.font = Font(name="Montserrat", size=11, bold=True, color=COLORS["white"])
        cell.fill = hex_fill(COLORS["green_dark"])
        cell.alignment = Alignment(horizontal="center")
    for row_idx, (habit, category) in enumerate(HABITS, start=2):
        ws.cell(row_idx, 1, row_idx - 1)
        ws.cell(row_idx, 2, habit)
        ws.cell(row_idx, 3, category)
        if row_idx % 2 == 0:
            apply_range_fill(ws, f"A{row_idx}:C{row_idx}", COLORS["gray_alt"])
    set_all_borders(ws, "A1:C31")
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 16


def build_tracker(ws) -> None:
    style_sheet_defaults(ws)
    ws.title = "Tracker"
    ws.freeze_panes = "C3"

    ws["A1"] = "Selected Month"
    ws["A1"].font = Font(name="Montserrat", bold=True, color=COLORS["text"])
    ws["D1"] = date(2026, 3, 1)
    ws["D1"].number_format = "mmmm yyyy"
    ws["F1"] = "Change D1 to switch months"
    ws["F1"].font = Font(name="Montserrat", italic=True, color=COLORS["teal"])

    headers = ["#", "Habit", "Category"] + list(range(1, 32)) + ["Progress %", "Longest Streak", "Current Streak", "Sparkline"]
    for idx, header in enumerate(headers, 1):
        cell = ws.cell(2, idx, header)
        cell.font = Font(name="Montserrat", size=10, bold=True, color=COLORS["white"])
        cell.fill = hex_fill(COLORS["green_dark"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx in range(3, 33):
        habit_idx = row_idx - 1
        ws.cell(row_idx, 1, habit_idx - 1)
        ws.cell(row_idx, 2, f"=IF(Habits!B{habit_idx}=\"\",\"\",Habits!B{habit_idx})")
        ws.cell(row_idx, 3, f"=IF(Habits!C{habit_idx}=\"\",\"\",Habits!C{habit_idx})")
        for col_idx in range(4, 35):
            ws.cell(row_idx, col_idx, False)
            ws.cell(row_idx, col_idx).alignment = Alignment(horizontal="center")
        ws.cell(row_idx, 35, f"=IFERROR(COUNTIF(D{row_idx}:INDEX($D{row_idx}:$AH{row_idx},1,Helper!$AA$2),TRUE)/Helper!$AA$2,0)")
        ws.cell(row_idx, 36, f"=IF(Helper!$AA$2=0,0,MAX(ARRAYFORMULA(FREQUENCY(IF(D{row_idx}:INDEX($D{row_idx}:$AH{row_idx},1,Helper!$AA$2)=TRUE,COLUMN(D{row_idx}:INDEX($D{row_idx}:$AH{row_idx},1,Helper!$AA$2))),IF(D{row_idx}:INDEX($D{row_idx}:$AH{row_idx},1,Helper!$AA$2)=FALSE,COLUMN(D{row_idx}:INDEX($D{row_idx}:$AH{row_idx},1,Helper!$AA$2)))))))")
        ws.cell(row_idx, 37, f"=IF(Helper!$AA$2=0,0,IF(INDEX($D{row_idx}:$AH{row_idx},1,Helper!$AA$2)=FALSE,0,IFERROR(Helper!$AA$2-LOOKUP(2,1/(D{row_idx}:INDEX($D{row_idx}:$AH{row_idx},1,Helper!$AA$2)=FALSE),COLUMN(D{row_idx}:INDEX($D{row_idx}:$AH{row_idx},1,Helper!$AA$2))-COLUMN(D{row_idx})+1),Helper!$AA$2)))")
        ws.cell(row_idx, 38, f'=IF(Helper!$AA$2=0,"",SPARKLINE(ARRAYFORMULA(N(D{row_idx}:INDEX($D{row_idx}:$AH{row_idx},1,Helper!$AA$2))),{{"charttype","column";"max",1;"color1","#{COLORS["green"]}";"empty","ignore"}}))')
        if row_idx % 2 == 1:
            apply_range_fill(ws, f"A{row_idx}:AL{row_idx}", COLORS["gray_alt"])

    for row in range(3, 33):
        ws.cell(row, 35).number_format = "0%"
        ws.cell(row, 38).alignment = Alignment(horizontal="center")

    set_all_borders(ws, "A2:AL32")
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 15
    for idx in range(4, 35):
        ws.column_dimensions[get_column_letter(idx)].width = 4.2
        ws.cell(2, idx).number_format = "d"
    for idx in range(35, 39):
        ws.column_dimensions[get_column_letter(idx)].width = 14

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24

    checkbox_fill_rule = FormulaRule(formula=["D3=TRUE"], fill=hex_fill(COLORS["green_soft"]))
    ws.conditional_formatting.add("D3:AH32", checkbox_fill_rule)
    ws.conditional_formatting.add("AI3:AI32", CellIsRule(operator="greaterThanOrEqual", formula=["0.8"], fill=hex_fill(COLORS["green_soft"])))
    ws.conditional_formatting.add("AI3:AI32", CellIsRule(operator="between", formula=["0.4", "0.8"], fill=hex_fill(COLORS["amber_soft"])))
    ws.conditional_formatting.add("AI3:AI32", CellIsRule(operator="lessThan", formula=["0.4"], fill=hex_fill(COLORS["red_soft"])))


def build_summary(ws) -> None:
    style_sheet_defaults(ws)
    ws.title = "Summary"
    ws.freeze_panes = "A2"

    headers = ["#", "Habit", "Category", "Progress %", "Longest Streak", "Current Streak", "30-Day Trend"]
    for idx, header in enumerate(headers, 1):
        cell = ws.cell(1, idx, header)
        cell.font = Font(name="Montserrat", size=11, bold=True, color=COLORS["white"])
        cell.fill = hex_fill(COLORS["green_dark"])
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for row_idx in range(2, 32):
        tracker_row = row_idx + 1
        ws.cell(row_idx, 1, row_idx - 1)
        ws.cell(row_idx, 2, f"=Tracker!B{tracker_row}")
        ws.cell(row_idx, 3, f"=Tracker!C{tracker_row}")
        ws.cell(row_idx, 4, f"=Tracker!AI{tracker_row}")
        ws.cell(row_idx, 5, f"=Tracker!AJ{tracker_row}")
        ws.cell(row_idx, 6, f"=Tracker!AK{tracker_row}")
        ws.cell(row_idx, 7, f'=IF(Helper!$AA$2=0,"",SPARKLINE(ARRAYFORMULA(N(Tracker!D{tracker_row}:INDEX(Tracker!D{tracker_row}:AH{tracker_row},1,Helper!$AA$2))),{{"charttype","column";"max",1;"color1","#{COLORS["green"]}";"empty","ignore"}}))')
        ws.cell(row_idx, 4).number_format = "0%"
        if row_idx % 2 == 0:
            apply_range_fill(ws, f"A{row_idx}:G{row_idx}", COLORS["gray_alt"])
    set_all_borders(ws, "A1:G31")
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 18
    ws.conditional_formatting.add("D2:D31", CellIsRule(operator="greaterThanOrEqual", formula=["0.8"], fill=hex_fill(COLORS["green_soft"])))
    ws.conditional_formatting.add("D2:D31", CellIsRule(operator="between", formula=["0.4", "0.8"], fill=hex_fill(COLORS["amber_soft"])))
    ws.conditional_formatting.add("D2:D31", CellIsRule(operator="lessThan", formula=["0.4"], fill=hex_fill(COLORS["red_soft"])))
    gold_side = Side(style="medium", color=COLORS["gold"])
    ws.conditional_formatting.add(
        "F2:F31",
        CellIsRule(
            operator="greaterThanOrEqual",
            formula=["7"],
            border=Border(left=gold_side, right=gold_side, top=gold_side, bottom=gold_side),
        ),
    )


def build_helper(ws) -> None:
    style_sheet_defaults(ws)
    ws.title = "Helper"
    headers = {
        "A1": "Day",
        "B1": "Date",
        "C1": "Daily Completion %",
        "D1": "7-Day Rolling Avg",
        "E1": "Completed Count",
        "F1": "Habit Count",
        "G1": "Habit",
        "H1": "Completion %",
        "I1": "Category",
        "J1": "Current Streak",
        "K1": "Longest Streak",
        "L1": "Previous Month %",
        "M1": "Improvement",
        "N1": "Top N",
        "O1": "Top Habit",
        "P1": "Top Completion %",
        "Q1": "Label",
        "R1": "Streak Habit",
        "S1": "Longest",
        "T1": "Current",
        "U1": "Category",
        "V1": "Avg Completion %",
        "W1": "Progress Slice",
        "X1": "Value",
        "Y1": "Previous Month Start",
        "Z1": "Heatmap Start",
        "AA1": "Active Days",
        "AB1": "Days In Month",
    }
    for cell, value in headers.items():
        ws[cell] = value
        ws[cell].font = Font(name="Montserrat", size=10, bold=True, color=COLORS["white"])
        ws[cell].fill = hex_fill(COLORS["green_dark"])
        ws[cell].alignment = Alignment(horizontal="center")

    ws["Y2"] = '=EOMONTH(Tracker!$D$1,-1)+1'
    ws["Z2"] = '=DATE(YEAR(Tracker!$D$1),MONTH(Tracker!$D$1),1)-WEEKDAY(DATE(YEAR(Tracker!$D$1),MONTH(Tracker!$D$1),1),2)+1'
    ws["AB2"] = '=DAY(EOMONTH(Tracker!$D$1,0))'
    ws["AA2"] = '=IF(Tracker!$D$1="",0,IF(DATE(YEAR(TODAY()),MONTH(TODAY()),1)<DATE(YEAR(Tracker!$D$1),MONTH(Tracker!$D$1),1),0,IF(DATE(YEAR(TODAY()),MONTH(TODAY()),1)=DATE(YEAR(Tracker!$D$1),MONTH(Tracker!$D$1),1),DAY(TODAY()),DAY(EOMONTH(Tracker!$D$1,0)))))'
    ws["N2"] = "=Dashboard!$Q$11"
    ws["W2"] = "Completed"
    ws["W3"] = "Remaining"
    ws["X2"] = "=IFERROR(AVERAGE(H2:H31),0)"
    ws["X3"] = "=MAX(0,1-X2)"
    ws["X2"].number_format = ws["X3"].number_format = "0%"

    for row_idx in range(2, 33):
        day = row_idx - 1
        ws.cell(row_idx, 1, day)
        ws.cell(row_idx, 2, f'=IF(A{row_idx}>$AB$2,"",DATE(YEAR(Tracker!$D$1),MONTH(Tracker!$D$1),A{row_idx}))')
        ws.cell(row_idx, 3, f'=IF(B{row_idx}="","",COUNTIF(INDEX(Tracker!$D$3:$AH$32,0,A{row_idx}),TRUE)/MAX(1,COUNTA(Tracker!$B$3:$B$32)))')
        start_row = max(2, row_idx - 6)
        ws.cell(row_idx, 4, f'=IF(C{row_idx}="","",AVERAGE(C{start_row}:C{row_idx}))')
        ws.cell(row_idx, 5, f'=IF(B{row_idx}="","",COUNTIF(INDEX(Tracker!$D$3:$AH$32,0,A{row_idx}),TRUE))')
        ws.cell(row_idx, 6, '=COUNTA(Tracker!$B$3:$B$32)')
        ws.cell(row_idx, 3).number_format = "0%"
        ws.cell(row_idx, 4).number_format = "0%"
        ws.cell(row_idx, 2).number_format = "d-mmm"

    for row_idx in range(2, 32):
        summary_row = row_idx
        ws.cell(row_idx, 7, f"=Summary!B{summary_row}")
        ws.cell(row_idx, 8, f"=Summary!D{summary_row}")
        ws.cell(row_idx, 9, f"=Summary!C{summary_row}")
        ws.cell(row_idx, 10, f"=Summary!F{summary_row}")
        ws.cell(row_idx, 11, f"=Summary!E{summary_row}")
        ws.cell(row_idx, 12, f'=IFERROR(INDEX(FILTER(History!$D$2:$D,History!$A$2:$A=$Y$2,History!$B$2:$B=G{row_idx}),1),"")')
        ws.cell(row_idx, 13, f'=IF(L{row_idx}="","",H{row_idx}-L{row_idx})')
        ws.cell(row_idx, 8).number_format = "0%"
        ws.cell(row_idx, 12).number_format = "0%"
        ws.cell(row_idx, 13).number_format = "+0%;-0%"

    for row_idx in range(2, 22):
        rank = row_idx - 1
        ws.cell(row_idx, 15, f'=IF(ROWS($O$2:O{row_idx})>$N$2,"",INDEX($G$2:$G$31,MATCH(LARGE($H$2:$H$31,ROWS($O$2:O{row_idx})),$H$2:$H$31,0)))')
        ws.cell(row_idx, 16, f'=IF(O{row_idx}="","",LARGE($H$2:$H$31,ROWS($P$2:P{row_idx})))')
        ws.cell(row_idx, 17, f'=IF(P{row_idx}="","",TEXT(P{row_idx},"0%"))')
        ws.cell(row_idx, 16).number_format = "0%"

    for row_idx in range(2, 12):
        ws.cell(row_idx, 18, f'=IFERROR(INDEX($G$2:$G$31,MATCH(LARGE($J$2:$J$31,ROWS($R$2:R{row_idx})),$J$2:$J$31,0)),"")')
        ws.cell(row_idx, 19, f'=IF(R{row_idx}="","",INDEX($K$2:$K$31,MATCH(R{row_idx},$G$2:$G$31,0)))')
        ws.cell(row_idx, 20, f'=IF(R{row_idx}="","",LARGE($J$2:$J$31,ROWS($T$2:T{row_idx})))')

    for row_idx in range(2, 11):
        ws.cell(row_idx, 21, f'=IFERROR(INDEX(SORT(UNIQUE(FILTER(Habits!$C$2:$C$31,Habits!$C$2:$C$31<>""))),ROWS($U$2:U{row_idx})),"")')
        ws.cell(row_idx, 22, f'=IF(U{row_idx}="","",AVERAGEIF($I$2:$I$31,U{row_idx},$H$2:$H$31))')
        ws.cell(row_idx, 22).number_format = "0%"

    set_all_borders(ws, "A1:AB32")
    for col_idx in range(1, 29):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14
    ws.sheet_state = "hidden"


def build_history(ws) -> None:
    style_sheet_defaults(ws)
    ws.title = "History"
    headers = ["Month Start", "Habit", "Category", "Completion %", "Longest Streak", "Current Streak", "Archived At"]
    for idx, header in enumerate(headers, 1):
        cell = ws.cell(1, idx, header)
        cell.font = Font(name="Montserrat", size=11, bold=True, color=COLORS["white"])
        cell.fill = hex_fill(COLORS["green_dark"])
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 22


def style_tile(ws, start: str, end: str, accent: str, label: str, value_formula: str, subtext_formula: str | None = None, percent: bool = False) -> None:
    min_col = ws[start].column
    max_col = ws[end].column
    min_row = ws[start].row
    max_row = ws[end].row
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row, col)
            cell.fill = hex_fill(COLORS["white"])
            cell.border = Border(
                left=Side(style="medium", color=accent),
                right=Side(style="medium", color=accent),
                top=Side(style="medium", color=accent),
                bottom=Side(style="medium", color=accent),
            )
    label_cell = ws[start]
    label_cell.value = label
    label_cell.font = Font(name="Montserrat", size=9, bold=True, color=COLORS["teal"])
    label_cell.alignment = Alignment(horizontal="left", vertical="top")

    value_cell = ws.cell(min_row + 1, min_col)
    value_cell.value = value_formula
    value_cell.font = Font(name="Montserrat", size=18, bold=True, color=COLORS["text"])
    value_cell.alignment = Alignment(horizontal="left", vertical="center")
    if percent:
        value_cell.number_format = "0%"

    if subtext_formula:
        subtext_cell = ws.cell(min_row + 3, min_col)
        subtext_cell.value = subtext_formula
        subtext_cell.font = Font(name="Montserrat", size=9, italic=True, color=COLORS["text"])
        subtext_cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)


def build_dashboard(ws) -> None:
    style_sheet_defaults(ws)
    ws.title = "Dashboard"
    ws.freeze_panes = "A2"

    style_title(ws, "A1", '="Monthly Habit Dashboard — "&TEXT(Tracker!D1,"MMMM YYYY")', "Q2")
    ws["A3"] = "Clean monthly snapshot with live formulas, chart-ready helper data, and Google Sheets automation support."
    ws["A3"].font = Font(name="Montserrat", size=10, color=COLORS["teal"])

    for col in range(1, 18):
        ws.column_dimensions[get_column_letter(col)].width = 12
    for col in range(21, 28):
        ws.column_dimensions[get_column_letter(col)].hidden = True

    style_tile(ws, "A4", "C8", COLORS["green"], "TOTAL HABITS", "=COUNTA(Habits!B2:B31)", '="tracked this month"')
    style_tile(ws, "D4", "F8", COLORS["green"], "AVERAGE COMPLETION %", "=IFERROR(AVERAGE(Summary!D2:D31),0)", '="across active days"', percent=True)
    style_tile(ws, "G4", "I8", COLORS["amber"], "HABITS > 80%", '=COUNTIF(Summary!D2:D31,">=0.8")', '="high consistency habits"')
    style_tile(ws, "J4", "L8", COLORS["green_dark"], "ACTIVE STREAK LEADER", "=MAX(Summary!F2:F31)", '=IFERROR(INDEX(Summary!B2:B31,MATCH(MAX(Summary!F2:F31),Summary!F2:F31,0)),"—")')
    style_tile(ws, "M4", "O8", COLORS["amber"], "BEST IMPROVEMENT VS PREV", "=IFERROR(MAX(Helper!M2:M31),0)", '=IFERROR(INDEX(Helper!G2:G31,MATCH(MAX(Helper!M2:M31),Helper!M2:M31,0)),"Archive a month to activate")', percent=True)

    ws["Q11"] = 10
    dv = DataValidation(type="list", formula1='"5,10,20"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws["Q11"])

    ws["N10"] = "Top Habits"
    ws["N10"].font = Font(name="Montserrat", size=12, bold=True, color=COLORS["text"])
    ws["N11"] = "Show Top N"
    ws["Q11"].font = Font(name="Montserrat", size=11, bold=True, color=COLORS["text"])
    ws["Q11"].alignment = Alignment(horizontal="center")
    ws["N12"] = "Quick Actions"
    ws["N13"] = "1. Change month in Tracker!D1"
    ws["N14"] = "2. Run Tracker Tools > Setup / Refresh dashboard"
    ws["N15"] = "3. Run Archive current month before rolling forward"
    for r in range(10, 19):
        apply_range_fill(ws, f"N{r}:Q{r}", COLORS["gray_alt"] if r % 2 == 0 else COLORS["white"])
    set_all_borders(ws, "N10:Q18")

    ws["N17"] = "Rank"
    ws["O17"] = "Habit"
    ws["P17"] = "Completion"
    ws["Q17"] = "Label"
    for cell in ["N17", "O17", "P17", "Q17"]:
        ws[cell].font = Font(name="Montserrat", size=9, bold=True, color=COLORS["white"])
        ws[cell].fill = hex_fill(COLORS["green_dark"])

    for row_idx in range(18, 28):
        helper_row = row_idx - 16
        ws.cell(row_idx, 14, helper_row - 1)
        ws.cell(row_idx, 15, f"=Helper!O{helper_row}")
        ws.cell(row_idx, 16, f"=Helper!P{helper_row}")
        ws.cell(row_idx, 17, f"=Helper!Q{helper_row}")
        ws.cell(row_idx, 16).number_format = "0%"

    ws["A24"] = "Monthly Heatmap"
    ws["A24"].font = Font(name="Montserrat", size=12, bold=True, color=COLORS["text"])
    weekdays = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    for idx, label in enumerate(weekdays, start=1):
        cell = ws.cell(25, idx, label)
        cell.font = Font(name="Montserrat", size=9, bold=True, color=COLORS["white"])
        cell.fill = hex_fill(COLORS["green_dark"])
        cell.alignment = Alignment(horizontal="center")

    for row_idx in range(26, 32):
        for col_idx in range(1, 8):
            offset = (row_idx - 26) * 7 + (col_idx - 1)
            date_formula = f'=IF(MONTH(Helper!$Z$2+{offset})=MONTH(Tracker!$D$1),TEXT(Helper!$Z$2+{offset},"d")&CHAR(10)&TEXT(INDEX(Helper!$C$2:$C$32,DAY(Helper!$Z$2+{offset})),"0%"),"")'
            hidden_formula = f'=IF(MONTH(Helper!$Z$2+{offset})=MONTH(Tracker!$D$1),INDEX(Helper!$C$2:$C$32,DAY(Helper!$Z$2+{offset})),"")'
            ws.cell(row_idx, col_idx, date_formula)
            ws.cell(row_idx, col_idx).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row_idx, col_idx).font = Font(name="Montserrat", size=9, color=COLORS["text"])
            ws.cell(row_idx, col_idx + 20, hidden_formula)
            ws.cell(row_idx, col_idx + 20).number_format = "0%"
    set_all_borders(ws, "A25:G31")
    for row_idx in range(26, 32):
        ws.row_dimensions[row_idx].height = 34

    # Heatmap rules against hidden values U:AA.
    heatmap_range = "A26:G31"
    ws.conditional_formatting.add(heatmap_range, FormulaRule(formula=["U26=\"\""], fill=hex_fill(COLORS["gray_bg"])))
    ws.conditional_formatting.add(heatmap_range, FormulaRule(formula=["AND(U26<>\"\",U26<0.2)"], fill=hex_fill("F0F4F5")))
    ws.conditional_formatting.add(heatmap_range, FormulaRule(formula=["AND(U26>=0.2,U26<0.4)"], fill=hex_fill("DCEFE0")))
    ws.conditional_formatting.add(heatmap_range, FormulaRule(formula=["AND(U26>=0.4,U26<0.6)"], fill=hex_fill("C2E4C9")))
    ws.conditional_formatting.add(heatmap_range, FormulaRule(formula=["AND(U26>=0.6,U26<0.8)"], fill=hex_fill("96D2A2")))
    ws.conditional_formatting.add(heatmap_range, FormulaRule(formula=["U26>=0.8"], fill=hex_fill(COLORS["green"])))

    ws["A33"] = "Legend"
    ws["A34"] = "Success"
    ws["B34"] = "#6BBF77"
    ws["A35"] = "Warning"
    ws["B35"] = "#D9A441"
    ws["A36"] = "Background"
    ws["B36"] = "#EEF2F4"
    apply_range_fill(ws, "B34", COLORS["green"])
    apply_range_fill(ws, "B35", COLORS["amber"])
    apply_range_fill(ws, "B36", COLORS["gray_bg"])
    set_all_borders(ws, "A33:B36")

    ws.conditional_formatting.add("D5", CellIsRule(operator="greaterThanOrEqual", formula=["0.8"], fill=hex_fill(COLORS["green_soft"])))
    ws.conditional_formatting.add("D5", CellIsRule(operator="between", formula=["0.4", "0.8"], fill=hex_fill(COLORS["amber_soft"])))
    ws.conditional_formatting.add("D5", CellIsRule(operator="lessThan", formula=["0.4"], fill=hex_fill(COLORS["red_soft"])))


def create_charts(ws) -> None:
    line = LineChart()
    line.title = "Monthly Completion Trend"
    line.style = 13
    line.y_axis.title = "Completion %"
    line.x_axis.title = "Date"
    line.y_axis.scaling.min = 0
    line.y_axis.scaling.max = 1
    line.smooth = True
    line.height = 7
    line.width = 11
    data = Reference(ws.parent["Helper"], min_col=3, max_col=4, min_row=1, max_row=32)
    cats = Reference(ws.parent["Helper"], min_col=2, min_row=2, max_row=32)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    line.legend.position = "b"
    line.ser[0].graphicalProperties.line.solidFill = COLORS["green_dark"]
    line.ser[0].graphicalProperties.line.width = 28000
    line.ser[1].graphicalProperties.line.solidFill = COLORS["teal"]
    line.ser[1].graphicalProperties.line.width = 18000
    ws.add_chart(line, "A10")

    bar = BarChart()
    bar.type = "bar"
    bar.style = 10
    bar.title = "Habit Consistency (Top N)"
    bar.y_axis.title = "Habit"
    bar.x_axis.title = "Completion %"
    bar.height = 8
    bar.width = 11
    bar.legend = None
    bar_data = Reference(ws.parent["Helper"], min_col=16, max_col=16, min_row=1, max_row=21)
    bar_cats = Reference(ws.parent["Helper"], min_col=15, min_row=2, max_row=21)
    bar.add_data(bar_data, titles_from_data=True)
    bar.set_categories(bar_cats)
    bar.ser[0].graphicalProperties.solidFill = COLORS["green"]
    bar.ser[0].graphicalProperties.line.solidFill = COLORS["green_dark"]
    bar.dLbls = DataLabelList()
    bar.dLbls.showVal = True
    ws.add_chart(bar, "A38")

    streak = BarChart()
    streak.type = "col"
    streak.style = 12
    streak.title = "Streak Leaderboard"
    streak.y_axis.title = "Days"
    streak.x_axis.title = "Habit"
    streak.height = 8
    streak.width = 8
    streak_data = Reference(ws.parent["Helper"], min_col=19, max_col=20, min_row=1, max_row=11)
    streak_cats = Reference(ws.parent["Helper"], min_col=18, min_row=2, max_row=11)
    streak.add_data(streak_data, titles_from_data=True)
    streak.set_categories(streak_cats)
    streak.ser[0].graphicalProperties.solidFill = COLORS["green_dark"]
    streak.ser[1].graphicalProperties.solidFill = COLORS["amber"]
    streak.legend.position = "b"
    ws.add_chart(streak, "I24")

    radar = RadarChart()
    radar.type = "filled"
    radar.style = 26
    radar.title = "Category Radar"
    radar.height = 7
    radar.width = 8
    radar_data = Reference(ws.parent["Helper"], min_col=22, max_col=22, min_row=1, max_row=10)
    radar_cats = Reference(ws.parent["Helper"], min_col=21, min_row=2, max_row=10)
    radar.add_data(radar_data, titles_from_data=True)
    radar.set_categories(radar_cats)
    radar.ser[0].graphicalProperties.solidFill = COLORS["teal"]
    radar.ser[0].graphicalProperties.line.solidFill = COLORS["green_dark"]
    ws.add_chart(radar, "I38")

    donut = DoughnutChart()
    donut.title = "Monthly Progress"
    donut.holeSize = 68
    donut.height = 6
    donut.width = 5
    donut_data = Reference(ws.parent["Helper"], min_col=24, max_col=24, min_row=1, max_row=3)
    donut_labels = Reference(ws.parent["Helper"], min_col=23, min_row=2, max_row=3)
    donut.add_data(donut_data, titles_from_data=True)
    donut.set_categories(donut_labels)
    donut.series[0].data_points = [
        DataPoint(idx=0),
        DataPoint(idx=1),
    ]
    donut.series[0].data_points[0].graphicalProperties.solidFill = COLORS["green_dark"]
    donut.series[0].data_points[1].graphicalProperties.solidFill = COLORS["gray_line"]
    donut.legend = None
    ws.add_chart(donut, "P3")


def build_script_sheet(ws) -> None:
    style_sheet_defaults(ws)
    ws.title = "AppScript"
    ws.column_dimensions["A"].width = 120
    ws["A1"] = "Paste this into Extensions > Apps Script in Google Sheets"
    ws["A1"].font = Font(name="Montserrat", size=12, bold=True, color=COLORS["text"])
    for idx, line in enumerate(APP_SCRIPT.splitlines(), start=3):
        ws.cell(idx, 1, line)
        ws.cell(idx, 1).font = Font(name="Courier New", size=10, color=COLORS["text"])


def build_how_to_use(ws) -> None:
    style_sheet_defaults(ws)
    ws.title = "HowToUse"
    ws.column_dimensions["A"].width = 110
    lines = [
        "HOW TO COPY & USE",
        "",
        "1. Upload SerenatasHabitTracker_Enhanced_Dashboard.xlsx to Google Drive and open it with Google Sheets.",
        "2. Go to Extensions > Apps Script and paste the full code from the AppScript tab or habit_dashboard.gs.",
        "3. Run setupHabitDashboard() once so the file inserts checkboxes, refreshes named ranges, and rebuilds the charts as Google Sheets objects.",
        "4. Change Tracker!D1 to any month-start date like 2026-04-01.",
        "5. Keep editing the Habits sheet if you want different habits or categories.",
        "6. Before you roll to a new month, run Tracker Tools > Archive current month. That powers the Best Improvement KPI.",
        "7. Mobile note: the dashboard is arranged as top KPIs, then trend, heatmap, and charts vertically so it remains readable on narrow screens.",
    ]
    for idx, line in enumerate(lines, start=1):
        ws.cell(idx, 1, line)
        ws.cell(idx, 1).font = Font(name="Montserrat", size=11 if idx == 1 else 10, bold=(idx == 1), color=COLORS["text"])


def build_workbook() -> Workbook:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    dashboard = wb.create_sheet("Dashboard")
    tracker = wb.create_sheet("Tracker")
    habits = wb.create_sheet("Habits")
    summary = wb.create_sheet("Summary")
    helper = wb.create_sheet("Helper")
    history = wb.create_sheet("History")
    script_sheet = wb.create_sheet("AppScript")
    how_to = wb.create_sheet("HowToUse")

    build_habits(habits)
    build_tracker(tracker)
    build_summary(summary)
    build_helper(helper)
    build_history(history)
    build_dashboard(dashboard)
    create_charts(dashboard)
    build_script_sheet(script_sheet)
    build_how_to_use(how_to)

    add_named_range(wb, "Tracker_Month", "'Tracker'!$D$1")
    add_named_range(wb, "Tracker_Dates", "'Tracker'!$D$2:$AH$2")
    add_named_range(wb, "Tracker_Habits", "'Tracker'!$B$3:$B$32")
    add_named_range(wb, "Tracker_Categories", "'Tracker'!$C$3:$C$32")
    add_named_range(wb, "Tracker_DataRange", "'Tracker'!$D$3:$AH$32")
    add_named_range(wb, "Summary_Progress", "'Summary'!$D$2:$D$31")
    add_named_range(wb, "Dashboard_TopN", "'Dashboard'!$Q$11")

    return wb


def main() -> None:
    wb = build_workbook()
    wb.save(OUTPUT_XLSX)
    APPS_SCRIPT_FILE.write_text(APP_SCRIPT)
    print(f"Created {OUTPUT_XLSX}")
    print(f"Wrote {APPS_SCRIPT_FILE}")


if __name__ == "__main__":
    main()
